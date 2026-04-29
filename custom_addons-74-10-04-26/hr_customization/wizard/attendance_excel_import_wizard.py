# -*- coding: utf-8 -*-

import base64
import openpyxl
import pytz

from io import BytesIO
from datetime import datetime, date

from odoo import models, fields
from odoo.exceptions import UserError, ValidationError


class AttendanceExcelImportWizard(models.TransientModel):
    _name = 'attendance.excel.import.wizard'
    _description = 'Attendance Excel Import Wizard'

    file = fields.Binary(string="Excel File", required=True)
    filename = fields.Char(string="Filename")

    # ---------------------------------------------------------
    # SESSION LOGIC (UNCHANGED)
    # ---------------------------------------------------------
    def _get_sessions(self, status):
        mapping = {
            'Present': ('present', 'present'),
            'Absent': ('absent', 'absent'),
            'PM': ('present', 'absent'),
            '1/2PR+1/2LP': ('present', 'absent'),
            '1/2LP+1/2PR': ('absent', 'present'),
        }
        return mapping.get((status or '').strip(), (False, False))

    # ---------------------------------------------------------
    # SAFE UTC DATETIME
    # ---------------------------------------------------------
    def _make_utc_datetime(self, att_date, time_value):
        user_tz = pytz.timezone(self.env.user.tz or 'UTC')

        if isinstance(time_value, datetime):
            local_dt = time_value
        else:
            h, m, *s = str(time_value).split(':')
            local_dt = datetime.combine(
                att_date,
                datetime.min.time().replace(
                    hour=int(h),
                    minute=int(m),
                    second=int(s[0]) if s else 0
                )
            )

        return user_tz.localize(local_dt).astimezone(pytz.UTC).replace(tzinfo=None)

    # ---------------------------------------------------------
    # CACHE LEAVE TYPES
    # ---------------------------------------------------------
    def _get_leave_types(self):
        LeaveType = self.env['hr.leave.type']
        return {
            'paid': LeaveType.search(
                [('work_entry_type_id.code', '=', 'LEAVE120')],
                limit=1
            ),
            'unpaid': LeaveType.search(
                [('work_entry_type_id.code', '=', 'LEAVE90')],
                limit=1
            ),
        }

    # ---------------------------------------------------------
    # CHECK 10-DAY BLOCK RULE (MAX 1 PAID LEAVE)
    # ---------------------------------------------------------
    def _is_paid_leave_allowed_in_block(self, employee, paid_leave_type, att_date):
        HrLeave = self.env['hr.leave']

        day = att_date.day
        if day <= 10:
            start, end = 1, 10
        elif day <= 20:
            start, end = 11, 20
        else:
            start, end = 21, 31

        leaves = HrLeave.search([
            ('employee_id', '=', employee.id),
            ('holiday_status_id', '=', paid_leave_type.id),
            ('request_date_from', '>=', att_date.replace(day=start)),
            ('request_date_to', '<=', att_date.replace(day=min(end, 28))),
            ('state', '!=', 'cancel'),
        ])

        used = sum(
            0.5 if l.request_unit_half else 1
            for l in leaves
        )

        return used < 1  # max 1 paid leave per block

    # ---------------------------------------------------------
    # CREATE LEAVE SAFELY (PAID → UNPAID FALLBACK)
    # ---------------------------------------------------------
    def _create_leave_safe(self, employee, leave_type, leave_days, att_date, unpaid_type):
        HrLeave = self.env['hr.leave']

        leave = HrLeave.create({
            'employee_id': employee.id,
            'holiday_status_id': leave_type.id,
            'request_date_from': att_date,
            'request_date_to': att_date,
            'request_unit_half': leave_days == 0.5,
        })

        try:
            leave.action_validate()
        except ValidationError:
            # Paid failed → delete & create unpaid
            leave.unlink()

            leave = HrLeave.create({
                'employee_id': employee.id,
                'holiday_status_id': unpaid_type.id,
                'request_date_from': att_date,
                'request_date_to': att_date,
                'request_unit_half': leave_days == 0.5,
            })
            leave.action_validate()

    # ---------------------------------------------------------
    # MAIN IMPORT
    # ---------------------------------------------------------
    def action_import(self):
        if not self.file:
            raise UserError("Please upload an Excel file.")

        try:
            workbook = openpyxl.load_workbook(
                BytesIO(base64.b64decode(self.file)),
                data_only=True
            )
            sheet = workbook.active
        except Exception:
            raise UserError("Invalid Excel file.")

        Employee = self.env['hr.employee']
        HrAttendance = self.env['hr.attendance']
        AttendanceExcel = self.env['attendance.excel']

        leave_types = self._get_leave_types()
        paid_type = leave_types['paid']
        unpaid_type = leave_types['unpaid']

        for row in sheet.iter_rows(min_row=2, values_only=True):

            if not row or not row[1]:
                continue

            employee = Employee.search(
                [('barcode', '=', str(row[1]).strip())],
                limit=1
            )
            if not employee:
                continue

            excel_date = row[9]
            if isinstance(excel_date, datetime):
                att_date = excel_date.date()
            elif isinstance(excel_date, date):
                att_date = excel_date
            else:
                att_date = datetime.strptime(str(excel_date), "%d-%b-%y").date()

            status = str(row[17] or '').strip()
            first, second = self._get_sessions(status)

            check_in = row[12]
            check_out = row[14]

            # ---------------- RAW IMPORT ----------------
            excel_rec = AttendanceExcel.create({
                'emp_code': row[1],
                'emp_name': employee.name,
                'department': row[3],
                'att_date': att_date,
                'check_in': str(check_in) if check_in else False,
                'check_out': str(check_out) if check_out else False,
                'total_time': str(row[16]) if row[16] else False,
                'status': status,
                'first_session': first,
                'second_session': second,
            })

            # ---------------- ATTENDANCE ----------------
            if check_in and check_out and (
                    first == 'present' or second == 'present'
            ):
                HrAttendance.create({
                    'employee_id': employee.id,
                    'check_in': excel_rec.schedule_check_in,
                    'check_out': excel_rec.schedule_check_out,
                    'schedule_check_in': self._make_utc_datetime(att_date, check_in),
                    'schedule_check_out': self._make_utc_datetime(att_date, check_out),
                    # 'check_in': self._make_utc_datetime(att_date, check_in),
                    # 'check_out': self._make_utc_datetime(att_date, check_out),
                    # 'schedule_check_in': excel_rec.schedule_check_in,
                    # 'schedule_check_out': excel_rec.schedule_check_out,
                })

            # PM → no leave
            if status == 'PM':
                continue

            # ---------------- LEAVE DAYS ----------------
            if first == second == 'absent':
                leave_days = 1
            elif first != second:
                leave_days = 0.5
            else:
                continue

            # ---------------- PAID / UNPAID DECISION ----------------
            use_paid = (
                (employee.leave_eligibility or '').lower() == 'yes'
                and self._is_paid_leave_allowed_in_block(employee, paid_type, att_date)
            )

            leave_type = paid_type if use_paid else unpaid_type

            self._create_leave_safe(
                employee,
                leave_type,
                leave_days,
                att_date,
                unpaid_type
            )

        return {'type': 'ir.actions.act_window_close'}






# import base64
# import openpyxl
# import pytz
#
# from io import BytesIO
# from datetime import datetime, date
#
# from odoo import models, fields
# from odoo.exceptions import UserError
#
#
# class AttendanceExcelImportWizard(models.TransientModel):
#     _name = 'attendance.excel.import.wizard'
#     _description = 'Attendance Excel Import Wizard'
#
#     file = fields.Binary(string="Excel File", required=True)
#     filename = fields.Char(string="Filename")
#
#     # -----------------------------
#     # SESSION LOGIC FROM STATUS
#     # -----------------------------
#     def _get_sessions(self, status):
#         status = (status or '').strip()
#
#         if status == 'Present':
#             return 'present', 'present'
#         if status == 'Absent':
#             return 'absent', 'absent'
#         if status == 'PM':
#             return 'present', 'absent'
#         if status == '1/2PR+1/2LP':
#             return 'present', 'absent'
#         if status == '1/2LP+1/2PR':
#             return 'absent', 'present'
#
#         return False, False
#
#     # -----------------------------
#     # SAFE UTC DATETIME
#     # -----------------------------
#     def _make_utc_datetime(self, att_date, time_value):
#         user_tz = pytz.timezone(self.env.user.tz or 'UTC')
#
#         if isinstance(time_value, datetime):
#             local_dt = time_value
#         else:
#             h, m, *s = str(time_value).split(':')
#             local_dt = datetime.combine(
#                 att_date,
#                 datetime.min.time().replace(
#                     hour=int(h),
#                     minute=int(m),
#                     second=int(s[0]) if s else 0
#                 )
#             )
#
#         return user_tz.localize(local_dt).astimezone(pytz.UTC).replace(tzinfo=None)
#
#     # -----------------------------
#     # DETERMINE LEAVE TYPE WITH 10-DAY BLOCK RULE
#     # -----------------------------
#     def _get_leave_type(self, employee, leave_days=1, att_date=None):
#         LeaveType = self.env['hr.leave.type']
#         HrLeave = self.env['hr.leave'].sudo()
#
#         unpaid_leave = LeaveType.search([('work_entry_type_id.code', '=', 'LEAVE90')], limit=1)
#
#         if not (employee.leave_eligibility or '').lower() == 'yes':
#             return unpaid_leave
#
#         paid_leave = LeaveType.search([('work_entry_type_id.code', '=', 'LEAVE120')], limit=1)
#
#         att_date = att_date or date.today()
#         day_of_month = att_date.day
#
#         if 1 <= day_of_month <= 10:
#             block_start, block_end = 1, 10
#         elif 11 <= day_of_month <= 20:
#             block_start, block_end = 11, 20
#         else:
#             block_start, block_end = 21, 31
#
#         leaves_this_block = HrLeave.search([
#             ('employee_id', '=', employee.id),
#             ('request_date_from', '>=', att_date.replace(day=block_start)),
#             ('request_date_to', '<=', att_date.replace(day=min(block_end, 28))),
#             ('holiday_status_id.work_entry_type_id.code', '=', 'LEAVE120'),
#             ('state', '!=', 'cancel')
#         ])
#         leaves_count = sum(1 if not l.request_unit_half else 0.5 for l in leaves_this_block)
#
#         if leaves_count < 1:
#             return paid_leave
#         else:
#             return unpaid_leave
#
#     # -----------------------------
#     # MAIN IMPORT
#     # -----------------------------
#     def action_import(self):
#         if not self.file:
#             raise UserError("Please upload an Excel file.")
#
#         try:
#             workbook = openpyxl.load_workbook(
#                 BytesIO(base64.b64decode(self.file)),
#                 data_only=True
#             )
#             sheet = workbook.active
#         except Exception:
#             raise UserError("Invalid Excel file.")
#
#         AttendanceExcel = self.env['attendance.excel']
#         HrAttendance = self.env['hr.attendance']
#         HrLeave = self.env['hr.leave']
#         Employee = self.env['hr.employee']
#
#         for row_no, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
#
#             if not row or not row[1]:
#                 continue
#
#             emp_code = str(row[1]).strip()
#             employee = Employee.search([('barcode', '=', emp_code)], limit=1)
#             if not employee:
#                 continue
#
#             # Attendance date
#             excel_date = row[9]
#             if isinstance(excel_date, datetime):
#                 att_date = excel_date.date()
#             elif isinstance(excel_date, date):
#                 att_date = excel_date
#             else:
#                 att_date = datetime.strptime(str(excel_date), "%d-%b-%y").date()
#
#             status = str(row[17]).strip() if row[17] else ''
#             first_session, second_session = self._get_sessions(status)
#
#             check_in = row[12]
#             check_out = row[14]
#
#             # -----------------------------
#             # attendance.excel record
#             # -----------------------------
#             AttendanceExcel.create({
#                 'emp_code': emp_code,
#                 'emp_name': employee.name,
#                 'department': row[3] or False,
#                 'att_date': att_date,
#                 'check_in': str(check_in) if check_in else False,
#                 'check_out': str(check_out) if check_out else False,
#                 'total_time': str(row[16]) if row[16] else False,
#                 'status': status,
#                 'first_session': first_session,
#                 'second_session': second_session,
#             })
#
#             # -----------------------------
#             # PM CASE – only create attendance
#             # -----------------------------
#             if status == 'PM':
#                 if check_in and check_out:
#                     HrAttendance.create({
#                         'employee_id': employee.id,
#                         'check_in': self._make_utc_datetime(att_date, check_in),
#                         'check_out': self._make_utc_datetime(att_date, check_out),
#                     })
#                 continue  # skip leave creation
#
#             # -----------------------------
#             # HALF-DAY CASES
#             # -----------------------------
#             if status in ('1/2PR+1/2LP', '1/2LP+1/2PR'):
#                 if check_in and check_out:
#                     HrAttendance.create({
#                         'employee_id': employee.id,
#                         'check_in': self._make_utc_datetime(att_date, check_in),
#                         'check_out': self._make_utc_datetime(att_date, check_out),
#                     })
#
#                 leave_type = self._get_leave_type(employee, leave_days=0.5, att_date=att_date)
#                 leave_vals = {
#                     'employee_id': employee.id,
#                     'holiday_status_id': leave_type.id,
#                     'request_date_from': att_date,
#                     'request_date_to': att_date,
#                     'request_unit_half': True,
#                 }
#                 leave = HrLeave.with_context(
#                     mail_create_nosubscribe=True,
#                     mail_notrack=True
#                 ).create(leave_vals)
#                 leave.sudo().action_validate()
#                 continue
#
#             # -----------------------------
#             # FULL DAY PRESENT
#             # -----------------------------
#             if first_session == 'present' and second_session == 'present' and check_in and check_out:
#                 HrAttendance.create({
#                     'employee_id': employee.id,
#                     'check_in': self._make_utc_datetime(att_date, check_in),
#                     'check_out': self._make_utc_datetime(att_date, check_out),
#                 })
#                 continue
#
#             # -----------------------------
#             # FULL / HALF DAY ABSENT
#             # -----------------------------
#             if first_session == 'absent' and second_session == 'absent':
#                 leave_days = 1
#             elif first_session != second_session:
#                 leave_days = 0.5
#             else:
#                 continue
#
#             leave_type = self._get_leave_type(employee, leave_days=leave_days, att_date=att_date)
#             leave_vals = {
#                 'employee_id': employee.id,
#                 'holiday_status_id': leave_type.id,
#                 'request_date_from': att_date,
#                 'request_date_to': att_date,
#                 'request_unit_half': leave_days == 0.5,
#             }
#
#             leave = HrLeave.with_context(
#                 mail_create_nosubscribe=True,
#                 mail_notrack=True
#             ).create(leave_vals)
#             leave.sudo().action_validate()
#
#         return {'type': 'ir.actions.act_window_close'}
#
#
#
#
#
# # import base64
# # import openpyxl
# # import pytz
# #
# # from io import BytesIO
# # from datetime import datetime, date
# #
# # from odoo import models, fields
# # from odoo.exceptions import UserError
# #
# #
# # class AttendanceExcelImportWizard(models.TransientModel):
# #     _name = 'attendance.excel.import.wizard'
# #     _description = 'Attendance Excel Import Wizard'
# #
# #     file = fields.Binary(string="Excel File", required=True)
# #     filename = fields.Char(string="Filename")
# #
# #     # --------------------------------------------------
# #     # SESSION LOGIC FROM STATUS
# #     # --------------------------------------------------
# #     def _get_sessions(self, status):
# #         status = (status or '').strip()
# #
# #         if status == 'Present':
# #             return 'present', 'present'
# #         if status == 'Absent':
# #             return 'absent', 'absent'
# #         if status == 'PM':
# #             return 'present', 'absent'
# #         if status == '1/2PR+1/2LP':
# #             return 'present', 'absent'
# #         if status == '1/2LP+1/2PR':
# #             return 'absent', 'present'
# #
# #         return False, False
# #
# #     # --------------------------------------------------
# #     # SAFE UTC DATETIME
# #     # --------------------------------------------------
# #     def _make_utc_datetime(self, att_date, time_value):
# #         user_tz = pytz.timezone(self.env.user.tz or 'UTC')
# #
# #         if isinstance(time_value, datetime):
# #             local_dt = time_value
# #         else:
# #             h, m, *s = str(time_value).split(':')
# #             local_dt = datetime.combine(
# #                 att_date,
# #                 datetime.min.time().replace(
# #                     hour=int(h),
# #                     minute=int(m),
# #                     second=int(s[0]) if s else 0
# #                 )
# #             )
# #
# #         return user_tz.localize(local_dt).astimezone(pytz.UTC).replace(tzinfo=None)
# #
# #     # --------------------------------------------------
# #     # DETERMINE LEAVE TYPE WITH 10-DAY BLOCK PAID/UNPAID RULE
# #     # --------------------------------------------------
# #     def _get_leave_type(self, employee, leave_days=1, att_date=None):
# #         LeaveType = self.env['hr.leave.type']
# #         HrLeave = self.env['hr.leave'].sudo()
# #
# #         # Unpaid leave type
# #         unpaid_leave = LeaveType.search([('work_entry_type_id.code', '=', 'LEAVE90')], limit=1)
# #
# #         # Not eligible for paid leave
# #         if not (employee.leave_eligibility or '').lower() == 'yes':
# #             return unpaid_leave
# #
# #         # Paid leave type
# #         paid_leave = LeaveType.search([('work_entry_type_id.code', '=', 'LEAVE120')], limit=1)
# #
# #         att_date = att_date or date.today()
# #         day_of_month = att_date.day
# #
# #         # Determine 10-day block
# #         if 1 <= day_of_month <= 10:
# #             block_start = 1
# #             block_end = 10
# #         elif 11 <= day_of_month <= 20:
# #             block_start = 11
# #             block_end = 20
# #         else:
# #             block_start = 21
# #             block_end = 31
# #
# #         # Count paid leaves already taken in this block
# #         leaves_this_block = HrLeave.search([
# #             ('employee_id', '=', employee.id),
# #             ('request_date_from', '>=', att_date.replace(day=block_start)),
# #             ('request_date_to', '<=', att_date.replace(day=min(block_end, 28))),  # safe last day
# #             ('holiday_status_id.work_entry_type_id.code', '=', 'LEAVE120'),
# #             ('state', '!=', 'cancel')
# #         ])
# #         leaves_count = sum(1 if not l.request_unit_half else 0.5 for l in leaves_this_block)
# #
# #         # Max 1 paid leave per 10-day block
# #         if leaves_count < 1:
# #             return paid_leave
# #         else:
# #             return unpaid_leave
# #
# #     # --------------------------------------------------
# #     # MAIN IMPORT
# #     # --------------------------------------------------
# #     def action_import(self):
# #         if not self.file:
# #             raise UserError("Please upload an Excel file.")
# #
# #         try:
# #             workbook = openpyxl.load_workbook(
# #                 BytesIO(base64.b64decode(self.file)),
# #                 data_only=True
# #             )
# #             sheet = workbook.active
# #         except Exception:
# #             raise UserError("Invalid Excel file.")
# #
# #         AttendanceExcel = self.env['attendance.excel']
# #         HrAttendance = self.env['hr.attendance']
# #         HrLeave = self.env['hr.leave']
# #         Employee = self.env['hr.employee']
# #
# #         for row_no, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
# #
# #             if not row or not row[1]:
# #                 continue
# #
# #             emp_code = str(row[1]).strip()
# #             employee = Employee.search([('barcode', '=', emp_code)], limit=1)
# #             if not employee:
# #                 continue
# #
# #             # Attendance date
# #             excel_date = row[9]
# #             if isinstance(excel_date, datetime):
# #                 att_date = excel_date.date()
# #             elif isinstance(excel_date, date):
# #                 att_date = excel_date
# #             else:
# #                 att_date = datetime.strptime(str(excel_date), "%d-%b-%y").date()
# #
# #             status = str(row[17]).strip() if row[17] else ''
# #             first_session, second_session = self._get_sessions(status)
# #
# #             check_in = row[12]
# #             check_out = row[14]
# #
# #             # --------------------------------------------------
# #             # attendance.excel record
# #             # --------------------------------------------------
# #             AttendanceExcel.create({
# #                 'emp_code': emp_code,
# #                 'emp_name': employee.name,
# #                 'department': row[3] or False,
# #                 'att_date': att_date,
# #                 'check_in': str(check_in) if check_in else False,
# #                 'check_out': str(check_out) if check_out else False,
# #                 'total_time': str(row[16]) if row[16] else False,
# #                 'status': status,
# #                 'first_session': first_session,
# #                 'second_session': second_session,
# #             })
# #
# #             # --------------------------------------------------
# #             # PM CASE – only create attendance
# #             # --------------------------------------------------
# #             if status == 'PM':
# #                 if check_in and check_out:
# #                     HrAttendance.create({
# #                         'employee_id': employee.id,
# #                         'check_in': self._make_utc_datetime(att_date, check_in),
# #                         'check_out': self._make_utc_datetime(att_date, check_out),
# #                     })
# #                 continue  # skip leave creation completely
# #
# #             # --------------------------------------------------
# #             # HALF-DAY CASES 1/2PR+1/2LP OR 1/2LP+1/2PR
# #             # --------------------------------------------------
# #             if status in ('1/2PR+1/2LP', '1/2LP+1/2PR'):
# #                 # create attendance for present session
# #                 if check_in and check_out:
# #                     HrAttendance.create({
# #                         'employee_id': employee.id,
# #                         'check_in': self._make_utc_datetime(att_date, check_in),
# #                         'check_out': self._make_utc_datetime(att_date, check_out),
# #                     })
# #                 # create half-day leave for absent session
# #                 leave_type = self._get_leave_type(employee, leave_days=0.5, att_date=att_date)
# #                 leave_vals = {
# #                     'employee_id': employee.id,
# #                     'holiday_status_id': leave_type.id,
# #                     'request_date_from': att_date,
# #                     'request_date_to': att_date,
# #                     'request_unit_half': True,
# #                 }
# #                 leave = HrLeave.with_context(
# #                     skip_work_entries=True,
# #                     mail_create_nosubscribe=True,
# #                     mail_notrack=True
# #                 ).create(leave_vals)
# #                 leave.sudo().write({'state': 'validate'})
# #                 continue
# #
# #             # --------------------------------------------------
# #             # FULL DAY PRESENT
# #             # --------------------------------------------------
# #             if first_session == 'present' and second_session == 'present' and check_in and check_out:
# #                 HrAttendance.create({
# #                     'employee_id': employee.id,
# #                     'check_in': self._make_utc_datetime(att_date, check_in),
# #                     'check_out': self._make_utc_datetime(att_date, check_out),
# #                 })
# #                 continue
# #
# #             # --------------------------------------------------
# #             # FULL / HALF DAY ABSENT
# #             # --------------------------------------------------
# #             if first_session == 'absent' and second_session == 'absent':
# #                 leave_days = 1
# #             elif first_session != second_session:
# #                 leave_days = 0.5
# #             else:
# #                 continue
# #
# #             leave_type = self._get_leave_type(employee, leave_days=leave_days, att_date=att_date)
# #             leave_vals = {
# #                 'employee_id': employee.id,
# #                 'holiday_status_id': leave_type.id,
# #                 'request_date_from': att_date,
# #                 'request_date_to': att_date,
# #                 'request_unit_half': leave_days == 0.5,
# #             }
# #
# #             leave = HrLeave.with_context(
# #                 skip_work_entries=True,
# #                 mail_create_nosubscribe=True,
# #                 mail_notrack=True
# #             ).create(leave_vals)
# #             leave.sudo().write({'state': 'validate'})
# #
# #         return {'type': 'ir.actions.act_window_close'}
# #
# #
# #
